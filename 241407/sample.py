import openpyxl as xl

input_path = "sample.xlsx"
output_path = "output.xlsx"

wb = xl.load_workbook(input_path)

# Sheet1 の商品一覧を取得する。
sheet1_items = [row[0].value for row in wb["Sheet1"].iter_rows(min_row=2)]
print(sheet1_items)  # ['りんご', 'みかん']


def get_option_value(item_name):
    """指定した名前の商品のオプション金額をシート "Sheet3" から取得する。
    """
    for row in wb["Sheet3"].iter_rows(min_row=2):
        if row[0].value == item_name:
            return row[2].value

    return None


for row in wb["Sheet2"].iter_rows(min_row=2):
    if row[0].value not in sheet1_items:
        # シート1に含まれない場合は、シート3からオプション金額をとってくる。
        row[2].value = get_option_value(row[0].value)

wb.save(output_path)
